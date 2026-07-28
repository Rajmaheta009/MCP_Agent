from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any


def extract_text(path: str) -> str:
    file = Path(path)
    if not file.exists():
        raise FileNotFoundError(path)

    ext = file.suffix.lower()
    if ext in {".txt", ".md", ".py", ".json"}:
        return file.read_text(encoding="utf-8", errors="ignore")

    if ext == ".pdf":
        from pypdf import PdfReader
        reader = PdfReader(str(file))
        return "\n".join(page.extract_text() or "" for page in reader.pages)

    if ext == ".docx":
        from docx import Document
        doc = Document(str(file))
        return "\n".join(p.text for p in doc.paragraphs)

    if ext in {".csv"}:
        with file.open("r", encoding="utf-8", errors="ignore", newline="") as fh:
            return "\n".join(" | ".join(row) for row in csv.reader(fh))

    if ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(str(file), read_only=True, data_only=True)
        lines: list[str] = []
        for ws in wb.worksheets:
            lines.append(f"# Sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                lines.append(" | ".join("" if v is None else str(v) for v in row))
        return "\n".join(lines)

    if ext in {".jpg", ".jpeg", ".png", ".webp"}:
        from PIL import Image
        with Image.open(file) as image:
            return (
                f"Image file: {file.name}\n"
                f"Format: {image.format}\n"
                f"Size: {image.size[0]}x{image.size[1]}\n"
                "OCR is not enabled in this version."
            )

    raise ValueError(f"Unsupported file type: {ext}")


def load_json(path: str) -> Any:
    return json.loads(Path(path).read_text(encoding="utf-8"))
